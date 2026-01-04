#!/usr/bin/env python3
"""
summarize_budget.py

Reads an existing budget workbook (created by your ingest script) and builds a "Budget" sheet.

Budget sheet layout:
- Columns: Date, Account, Note, <one column per category>
- Row 1: headers
- Row 2: SUM row (sums each category column)
- Data rows start at row 3
- Copies every transaction from selected account sheets EXCEPT Category == "Ignore"
- Places the transaction Amount into the column matching its Category (one-hot by category)
- Sorts Budget rows by Category then Date (category order from CONFIG)
- Moves the Budget sheet to the FRONT (first tab)
- Allows manual/extra transactions specified in CONFIG to be appended into Budget
  * date defaults to the first day of the month if omitted:
      - If CONFIG['month'] is provided (YYYY-MM), uses that month
      - Else uses the month implied by CONFIG['excel_file'] timestamp? (not reliable)
      - Else uses the current month on your machine

Dependencies:
  pip install openpyxl
"""

from __future__ import annotations

import os
from datetime import datetime, date
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

import warnings

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
    category=UserWarning,
    module="openpyxl"
)



# ======================================================================
# 1) CONFIGURATION (edit this section only)
# ======================================================================

CONFIG: Dict[str, Any] = {

    # Workbook file (must exist)
    "excel_file": "12 Dec 2025.xlsx",

    # Optional: month context for default manual transaction dates, format "YYYY-MM"
    # If omitted, defaults to the current month on your machine.
    "month": "2025-12",

    # Which sheets to include (by sheet name in Excel).
    "account_sheets_to_include": [
        "CapitalOne Visa Quicksilver 766",
        "Cash",
        "Chase Checking Company 4272",
        "Chase Checking Farm 5356",
        "Chase Visa Company 2983",
        "Citi Visa Costco 0345",
        "MVB Checking 1194",
        "PayPal Checking Coy",
        "PayPal Checking Kiki",
        "Schwab Checking Joint 4446",
        "Schwab Checking Travel 7183",
        "Venmo Checking Coy",
        "Venmo Checking Kiki",
        "WellsFargo Visa ActiveCash 1758",
    ],

    # Categories (these become columns)
    "categories": [
        "Auto", "Donation", "Eating Out", "Groceries", "Gypsy", "Health", "Ignore",
        "Jasey", "Misc", "Shelter", "Tegan", "Vacation",
    ],

    "ignore_category_value": "Ignore",
    "budget_sheet_name": "Budget",

    # Optional: add manual transactions directly to Budget.
    # Each item:
    #   - account: "Some Account" (required; shown in Budget Account column)
    #   - category: must match one of CONFIG["categories"] (except Ignore) (required)
    #   - amount: number (required)
    #   - note: "text" (optional)
    #   - date: "YYYY-MM-DD" (optional; defaults to first day of CONFIG['month'])
    #
    # Example:
    # "manual_budget_transactions": [
    #     {"account": "Manual", "note": "Cash tip", "category": "Misc", "amount": -10.00},
    #     {"date": "2025-11-05", "account": "Manual", "note": "Reimbursement", "category": "Misc", "amount": 25.00},
    # ],
    "manual_budget_transactions": [
        {"account": "Schwab Checking Joint 4446", "note": "Car Insurance", "category": "Auto", "amount": -200},
        {"account": "Schwab Checking Joint 4446", "note": "House Insurance", "category": "Shelter", "amount": -43},
        {"account": "Schwab Checking Joint 4446", "note": "Trailer Insurance", "category": "Misc", "amount": -37.25},
        {"account": "Schwab Checking Joint 4446", "note": "Unbrella Insurance", "category": "Misc", "amount": -66.88},
        {"account": "Schwab Checking Joint 4446", "note": "Tegan Renters Insurance", "category": "Tegan", "amount": -15.58},
        {"account": "Schwab Checking Joint 4446", "note": "Trailer Parking", "category": "Misc", "amount": -150},
    ],

    # Formatting
    "wrap_text": True,
    "header_bold": True,
    "freeze_panes": "A3",

    # Column widths
    "width_date": 12,
    "width_account": 28,
    "width_note": 40,
    "width_category": 15,

    # Alignment (openpyxl expects: left/center/right/general/etc.)
    "align_date": "right",
    "align_account": "left",
    "align_note": "left",
    "align_category": "right",
}


# ======================================================================
# 2) BASIC HELPERS
# ======================================================================

def norm_text(v: Any) -> str:
    return ("" if v is None else str(v)).strip().lower()


def norm_category(v: Any) -> str:
    return " ".join(norm_text(v).split())


def safe_date_cell(v: Any) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    try:
        dt = datetime.fromisoformat(str(v))
        return dt.date()
    except Exception:
        return None


def parse_iso_date(s: str) -> date:
    try:
        return datetime.strptime(s.strip(), "%Y-%m-%d").date()
    except Exception as e:
        raise ValueError(f"Invalid date '{s}'. Expected YYYY-MM-DD") from e


def parse_month_first_day(month_str: str) -> date:
    try:
        dt = datetime.strptime(month_str.strip(), "%Y-%m")
        return date(dt.year, dt.month, 1)
    except Exception as e:
        raise ValueError(f"Invalid CONFIG['month'] value '{month_str}'. Expected 'YYYY-MM'") from e


def default_manual_date(cfg: Dict[str, Any]) -> date:
    m = str(cfg.get("month", "")).strip()
    if m != "":
        return parse_month_first_day(m)
    today = date.today()
    return date(today.year, today.month, 1)


def safe_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except Exception:
        return None


# ======================================================================
# 3) READ TRANSACTIONS FROM ACCOUNT SHEETS
# ======================================================================

STANDARD_HEADERS = ["Date", "Description", "Note", "Category", "Amount", "Balance"]

def locate_standard_columns(ws) -> Dict[str, int]:
    header_row = 1
    mapping: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        name = str(v).strip()
        if name in STANDARD_HEADERS:
            mapping[name] = c

    required = ["Date", "Note", "Category", "Amount"]
    missing = [h for h in required if h not in mapping]
    if missing:
        raise KeyError(f"Sheet '{ws.title}' missing required headers: {missing}. Found: {list(mapping.keys())}")

    return mapping


def iter_transactions(ws, col_map: Dict[str, int]):
    empty_streak = 0
    max_empty_streak = 50

    for r in range(2, ws.max_row + 1):
        dv = ws.cell(row=r, column=col_map["Date"]).value
        note = ws.cell(row=r, column=col_map["Note"]).value
        cat = ws.cell(row=r, column=col_map["Category"]).value
        amt = ws.cell(row=r, column=col_map["Amount"]).value

        d = safe_date_cell(dv)
        a = safe_float(amt)
        c = norm_category(cat)

        if (d is None) and (a is None) and (note in (None, "")) and (cat in (None, "")):
            empty_streak += 1
            if empty_streak >= max_empty_streak:
                break
            continue

        empty_streak = 0
        yield (d, note, c, a)


# ======================================================================
# 4) BUDGET SHEET: BUILD / SORT / WRITE (includes manual transactions)
# ======================================================================

def clear_sheet(ws) -> None:
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    if ws.max_column > 0:
        ws.delete_cols(1, ws.max_column)


def apply_budget_formatting(ws, *, cfg: Dict[str, Any], num_categories: int) -> None:
    wrap_text = bool(cfg.get("wrap_text", True))
    header_bold = bool(cfg.get("header_bold", True))

    widths = [
        cfg.get("width_date", 12),
        cfg.get("width_account", 28),
        cfg.get("width_note", 40),
    ] + [cfg.get("width_category", 15)] * num_categories

    aligns = [
        cfg.get("align_date", "right"),
        cfg.get("align_account", "left"),
        cfg.get("align_note", "left"),
    ] + [cfg.get("align_category", "right")] * num_categories

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    max_r = ws.max_row
    max_c = 3 + num_categories

    if header_bold:
        for c in range(1, max_c + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)
            ws.cell(row=2, column=c).font = Font(bold=True)

    for c in range(1, max_c + 1):
        align = Alignment(horizontal=aligns[c - 1], wrap_text=wrap_text)
        for r in range(1, max_r + 1):
            ws.cell(row=r, column=c).alignment = align

    freeze = cfg.get("freeze_panes", None)
    if freeze:
        ws.freeze_panes = freeze


def move_sheet_to_front(wb, sheet_name: str) -> None:
    if sheet_name not in wb.sheetnames:
        return
    ws_obj = wb[sheet_name]
    wb._sheets.remove(ws_obj)
    wb._sheets.insert(0, ws_obj)


def compute_row_category_from_budget_row(row_vals: List[Any], category_cols: List[str]) -> str:
    for i, cat_name in enumerate(category_cols):
        v = row_vals[3 + i]
        if v is None or v == "":
            continue
        try:
            _ = float(v)
            return cat_name
        except Exception:
            continue
    return ""


def sort_budget_data_rows(
    rows: List[List[Any]],
    *,
    category_cols: List[str],
    category_order: List[str]
) -> List[List[Any]]:
    order_map: Dict[str, int] = {}
    for i, c in enumerate(category_order):
        order_map[norm_category(c)] = i

    def key_fn(r: List[Any]) -> Tuple[int, date]:
        cat = compute_row_category_from_budget_row(r, category_cols)
        cat_idx = order_map.get(norm_category(cat), 10**9)

        d = safe_date_cell(r[0])
        if d is None:
            d = date(1900, 1, 1)

        return (cat_idx, d)

    return sorted(rows, key=key_fn)


def validate_manual_transactions(
    manual: List[Dict[str, Any]],
    *,
    cfg: Dict[str, Any],
    categories_for_columns: List[str],
    ignore_value: str
) -> List[Tuple[date, str, str, str, float]]:
    """
    Returns list of tuples: (date, account, note, category_name, amount)
    date defaults to first day of cfg['month'] if omitted.
    """
    cat_allowed = {norm_category(c): c for c in categories_for_columns}
    out_rows: List[Tuple[date, str, str, str, float]] = []
    default_d = default_manual_date(cfg)

    for i, t in enumerate(manual):
        if not isinstance(t, dict):
            raise ValueError(f"manual_budget_transactions[{i}] must be an object/dict")

        d_raw = t.get("date", None)
        acct = t.get("account", None)
        note = t.get("note", "")
        cat = t.get("category", None)
        amt = t.get("amount", None)

        if acct is None or str(acct).strip() == "":
            raise ValueError(f"manual_budget_transactions[{i}] missing 'account'")
        if cat is None or str(cat).strip() == "":
            raise ValueError(f"manual_budget_transactions[{i}] missing 'category'")
        if amt is None or str(amt).strip() == "":
            raise ValueError(f"manual_budget_transactions[{i}] missing 'amount'")

        d = default_d if (d_raw is None or str(d_raw).strip() == "") else parse_iso_date(str(d_raw))

        cat_norm = norm_category(cat)
        if cat_norm == ignore_value:
            continue
        if cat_norm not in cat_allowed:
            raise ValueError(
                f"manual_budget_transactions[{i}] category '{cat}' not in allowed categories {categories_for_columns}"
            )

        a = float(amt)
        out_rows.append((d, str(acct), "" if note is None else str(note), cat_allowed[cat_norm], a))

    return out_rows


def write_budget_sheet(
    wb,
    *,
    cfg: Dict[str, Any],
    selected_sheets: List[str],
    categories: List[str]
) -> None:
    budget_sheet_name = str(cfg.get("budget_sheet_name", "Budget")).strip() or "Budget"

    if budget_sheet_name in wb.sheetnames:
        ws_budget = wb[budget_sheet_name]
        clear_sheet(ws_budget)
    else:
        ws_budget = wb.create_sheet(budget_sheet_name)

    ignore_value = norm_category(cfg.get("ignore_category_value", "Ignore"))
    categories_for_columns = [c for c in categories if norm_category(c) != ignore_value]

    headers = ["Date", "Account", "Note"] + categories_for_columns

    data_rows: List[List[Any]] = []

    cat_to_offset: Dict[str, int] = {}
    for i, c in enumerate(categories_for_columns):
        cat_to_offset[norm_category(c)] = i

    # From account sheets
    for sheet_name in selected_sheets:
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Configured sheet not found in workbook: '{sheet_name}'")

        ws_src = wb[sheet_name]
        col_map = locate_standard_columns(ws_src)

        for (d, note, cat_norm, amt) in iter_transactions(ws_src, col_map):
            if cat_norm == "" or amt is None:
                continue
            if cat_norm == ignore_value:
                continue
            if cat_norm not in cat_to_offset:
                continue

            row: List[Any] = [d, sheet_name, "" if note is None else str(note)] + [""] * len(categories_for_columns)
            row[3 + cat_to_offset[cat_norm]] = float(amt)
            data_rows.append(row)

    # Manual transactions from config (date defaults to first day of month)
    manual_cfg = cfg.get("manual_budget_transactions", [])
    if manual_cfg:
        manual_rows = validate_manual_transactions(
            manual_cfg,
            cfg=cfg,
            categories_for_columns=categories_for_columns,
            ignore_value=ignore_value
        )

        for (d, acct, note, cat_name, amt) in manual_rows:
            cat_norm = norm_category(cat_name)
            row = [d, acct, note] + [""] * len(categories_for_columns)
            row[3 + cat_to_offset[cat_norm]] = float(amt)
            data_rows.append(row)

    # Sort by category then date
    data_rows = sort_budget_data_rows(
        data_rows,
        category_cols=categories_for_columns,
        category_order=categories_for_columns
    )

    # Write header + sum row placeholders
    ws_budget.append(headers)
    ws_budget.append(["TOTAL", "", ""] + [""] * len(categories_for_columns))

    # Write sorted data rows starting at row 3
    for row in data_rows:
        ws_budget.append(row)
        r = ws_budget.max_row
        for c in range(4, 4 + len(categories_for_columns)):
            v = ws_budget.cell(row=r, column=c).value
            if v is None or v == "":
                continue
            ws_budget.cell(row=r, column=c).number_format = "0.00"

    # Fill sum formulas in row 2 (summing rows 3..last)
    last_data_row = ws_budget.max_row
    sum_row = 2

    for col_idx in range(4, 4 + len(categories_for_columns)):
        col_letter = get_column_letter(col_idx)
        if last_data_row >= 3:
            ws_budget.cell(row=sum_row, column=col_idx).value = f"=SUM({col_letter}3:{col_letter}{last_data_row})"
        else:
            ws_budget.cell(row=sum_row, column=col_idx).value = 0
        ws_budget.cell(row=sum_row, column=col_idx).number_format = "0.00"

    apply_budget_formatting(ws_budget, cfg=cfg, num_categories=len(categories_for_columns))
    move_sheet_to_front(wb, budget_sheet_name)


# ======================================================================
# 5) MAIN
# ======================================================================

def main() -> None:
    cfg = CONFIG

    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_file = str(cfg.get("excel_file", "")).strip()
    if not excel_file:
        raise ValueError("CONFIG['excel_file'] must be set (e.g., 'November 2025.xlsx').")

    excel_path = excel_file
    if not os.path.isabs(excel_path):
        excel_path = os.path.join(script_dir, excel_file)

    if not os.path.isfile(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = load_workbook(excel_path)

    selected_sheets = list(cfg.get("account_sheets_to_include", []))
    if not selected_sheets:
        raise ValueError("CONFIG['account_sheets_to_include'] must list at least one sheet.")

    categories = list(cfg.get("categories", []))
    if not categories:
        raise ValueError("CONFIG['categories'] must be provided.")

    write_budget_sheet(
        wb,
        cfg=cfg,
        selected_sheets=selected_sheets,
        categories=categories
    )

    wb.save(excel_path)
    print(f"Updated workbook with '{cfg.get('budget_sheet_name', 'Budget')}' sheet: {excel_path}")


if __name__ == "__main__":
    main()
