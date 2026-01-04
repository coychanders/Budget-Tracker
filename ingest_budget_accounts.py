#!/usr/bin/env python3
"""
Budget Accounts CSV -> Excel Workbook (append mode)

What this script does:
- Looks for <workbook_name>.xlsx in the same folder as this script.
- If it exists: APPENDS only NEW transactions not already present in each sheet.
- If it doesn't exist: creates it and writes all transactions.
- Reads all .csv files in each subfolder of CONFIG["budget_accounts_folder"].
  The subfolder name must match a key in CONFIG["accounts_by_name"].

Sheet columns (always, in this order):
  Date, Description, Note, Category, Amount, Balance

Duplicate detection for append mode:
- DEDUP IGNORES NOTE.
- A transaction is considered a duplicate if (Date + Description + Amount) matches an existing row.
- ALSO dedups within the ingested CSV batch (across multiple files for the same account) before appending.

Important fixes included:
- Venmo ending balance is written onto the last REAL transaction row (not below it).
- Existing keys and appends only consider REAL transaction rows (ignores formatting-inflated ws.max_row).
- Appends write directly below the last REAL transaction row (not at row 1001+).
- Drops non-transaction rows that don't have a Date (prevents Venmo header/summary lines from being appended).

Formatting:
- Wrap text ON for all REAL rows (header through last transaction row).
- Column widths set to: 15, 50, 50, 15, 15, 15  (A..F)
- Column alignments set to: right, left, left, left, right, right  (A..F)

Dependencies:
  pip install pandas openpyxl
"""

from __future__ import annotations

import glob
import os
import re
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple, Union, Set

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ======================================================================
# 1) CONFIGURATION (edit this section only)
# ======================================================================

MatchValue = Union[str, int]

STANDARD_COLUMNS: List[str] = ["Date", "Description", "Note", "Category", "Amount", "Balance"]

CONFIG: Dict[str, Any] = {

    # **********  UPDATE BEFORE RUNNING **********

    "workbook_name": "12 Dec 2025",
    "date_range": {"start": "2025-12-01", "stop": "2025-12-31"},  # inclusive
    "budget_accounts_folder": "Budget Accounts",

    # **********  CATEGORIES **********

    "categories": [
        "Auto", "Donation", "Eating Out", "Groceries", "Health", "Ignore", "Jasey", "Misc", "Pet", "Shelter",
        "Tegan", "Vacation"
    ],

    "skip_unknown_account_folders": True,
    "category_validation_min_rows": 1000,

    # **********  ACCOUNTS **********

    "accounts_by_name": {

        # CapitalOne Visa
        "CapitalOne_Visa_Quicksilver_7660": {
            "sheet_name": "CapitalOne Visa Quicksilver 7660",
            "has_header": True,
            "skiprows": 0,
            "delimiter": None,

            "match": {
                "Date": "Transaction Date",
                "Description": "Description",
                "Withdraw": "Debit",
                "Deposit": "Credit",
            },

            "derive_amount_from_withdraw_deposit": True,
            "withdraw_is_negative": False,
            "dedup_include_balance": False,
            "balance_mode": "running_balance_excluding_ignore",
        },

        # Cash
        "Cash": {
            "sheet_name": "Cash",
            "has_header": True,
            "skiprows": 0,
            "delimiter": None,

            "match": {
                "Date": "Date",
                "Description": "Description",
                "Amount": "Amount",
            },

            "derive_amount_from_withdraw_deposit": False,
            "withdraw_is_negative": True,
            "dedup_include_balance": False,
            "balance_mode": "none",
        },

        # Chase Checking Company 4272
        "Chase_Checking_Company_4272": {
            "sheet_name": "Chase Checking Company 4272",
            "has_header": True,
            "skiprows": 0,
            "delimiter": None,

            "match": {
                "Date": "Posting Date",
                "Description": "Description",
                "Amount": "Amount",
                "Balance": "Balance",
            },

            "derive_amount_from_withdraw_deposit": False,
            "withdraw_is_negative": True,
            "dedup_include_balance": False,
            "balance_mode": "none",
        },

        # Chase Checking Farm 5356
        "Chase_Checking_Farm_5356": {
            "sheet_name": "Chase Checking Farm 5356",
            "has_header": True,
            "skiprows": 0,
            "delimiter": None,

            "match": {
                "Date": "Posting Date",
                "Description": "Description",
                "Amount": "Amount",
                "Balance": "Balance",
            },

            "derive_amount_from_withdraw_deposit": False,
            "withdraw_is_negative": True,
            "dedup_include_balance": False,
            "balance_mode": "none",
        },

        # Chase Visa Company 2983
        "Chase_Visa_Company_2983": {
            "sheet_name": "Chase Visa Company 2983",
            "has_header": True,
            "skiprows": 0,
            "delimiter": None,

            "match": {
                "Date": "Transaction Date",
                "Description": "Description",
                "Amount": "Amount",
            },

            "derive_amount_from_withdraw_deposit": False,
            "withdraw_is_negative": True,
            "dedup_include_balance": False,
            "balance_mode": "none",
        },

        # Citi Visa Costco 0345
        "Citi_Visa_Costco_0345": {
            "sheet_name": "Citi Visa Costco 0345",
            "has_header": True,
            "skiprows": 0,
            "delimiter": None,

            "match": {
                "Date": "Date",
                "Description": "Description",
                "Withdraw": "Debit",
                "Deposit": "Credit",
            },

            "derive_amount_from_withdraw_deposit": True,
            "withdraw_is_negative": False,
            "dedup_include_balance": False,
            "balance_mode": "running_balance_excluding_ignore",
        },

        # MVB Checking 1194
        "MVB_Checking_1194": {
            "sheet_name": "MVB Checking 1194",
            "has_header": True,
            "skiprows": 0,
            "delimiter": None,

            "match": {
                "Date": "Processed Date",
                "Description": "Description",
                "Amount": "Amount",
            },

            "derive_amount_from_withdraw_deposit": False,
            "withdraw_is_negative": True,
            "dedup_include_balance": False,
            "balance_mode": "running_balance_excluding_ignore",
        },

        # PayPal_Checking_Coy
        "PayPal_Checking_Coy": {
            "sheet_name": "PayPal Checking Coy",
            "has_header": True,
            "skiprows": 0,
            "delimiter": None,

            "match": {
                "Date": "Date",
                "Description": "Item Title",
                "Amount": "Amount",
                "Balance": "Balance",
            },

            "derive_amount_from_withdraw_deposit": False,
            "withdraw_is_negative": True,
            "dedup_include_balance": False,
            "balance_mode": "none",
        },

        # PayPal_Checking_Kiki
        "PayPal_Checking_Kiki": {
            "sheet_name": "PayPal Checking Kiki",
            "has_header": True,
            "skiprows": 0,
            "delimiter": None,

            "match": {
                "Date": "Date",
                "Description": "Item Title",
                "Amount": "Amount",
                "Balance": "Balance",
            },

            "derive_amount_from_withdraw_deposit": False,
            "withdraw_is_negative": True,
            "dedup_include_balance": False,
            "balance_mode": "none",
        },

        # Schwab Joint Checking 4446
        "Schwab_Checking_Joint_4446": {
            "sheet_name": "Schwab Checking Joint 4446",
            "has_header": True,
            "skiprows": 0,
            "delimiter": None,

            "match": {
                "Date": "Date",
                "Description": "Description",
                "Withdraw": "Withdrawal",
                "Deposit": "Deposit",
                "Balance": "RunningBalance",
            },

            "derive_amount_from_withdraw_deposit": True,
            "withdraw_is_negative": False,
            "dedup_include_balance": False,
            "balance_mode": "none",
        },

        # Schwab_Checking_Travel_7183
        "Schwab_Checking_Travel_7183": {
            "sheet_name": "Schwab Checking Travel 7183",
            "has_header": True,
            "skiprows": 0,
            "delimiter": None,

            "match": {
                "Date": "Date",
                "Description": "Description",
                "Withdraw": "Withdrawal",
                "Deposit": "Deposit",
                "Balance": "RunningBalance",
            },

            "derive_amount_from_withdraw_deposit": True,
            "withdraw_is_negative": False,
            "dedup_include_balance": False,
            "balance_mode": "none",
        },

        # Venmo Checking Coy
        "Venmo_Checking_Coy": {
            "sheet_name": "Venmo Checking Coy",
            "has_header": True,
            "delimiter": None,
            "header_search_tokens": ["ID", "Datetime", "Type", "Status", "Note"],
            "max_header_scan_lines": 300,
            "ending_balance_header": "Ending Balance",
            "date_header_for_balance_pick": "Datetime",
            "place_ending_balance_on_last_transaction_row": True,

            "match": {
                "Date": "Datetime",
                "Description": "Note",
                "Amount": "Amount (total)",
                "Balance": "Ending Balance",
            },

            "derive_amount_from_withdraw_deposit": False,
            "dedup_include_balance": False,
            "balance_mode": "none",
        },

        # Venmo Checking Kiki
        "Venmo_Checking_Kiki": {
            "sheet_name": "Venmo Checking Kiki",
            "has_header": True,
            "delimiter": None,
            "header_search_tokens": ["ID", "Datetime", "Type", "Status", "Note"],
            "max_header_scan_lines": 300,
            "ending_balance_header": "Ending Balance",
            "date_header_for_balance_pick": "Datetime",
            "place_ending_balance_on_last_transaction_row": True,

            "match": {
                "Date": "Datetime",
                "Description": "Note",
                "Amount": "Amount (total)",
                "Balance": "Ending Balance",
            },

            "derive_amount_from_withdraw_deposit": False,
            "dedup_include_balance": False,
            "balance_mode": "none",
        },

        # WellsFargo Visa ActiveCash 1758
        "WellsFargo_Visa_ActiveCash_1758": {
            "sheet_name": "WellsFargo Visa ActiveCash 1758",
            "has_header": False,
            "skiprows": 0,
            "delimiter": None,

            "match": {
                "Date": 0,
                "Amount": 1,
                "Description": 4,
            },

            "derive_amount_from_withdraw_deposit": False,
            "dedup_include_balance": False,
            "balance_mode": "running_balance_excluding_ignore",
        },
    },

    # **********  CATEGORY RULES **********
    # NOTE: Rules use ONLY the key "note"

    "auto_category_rules": {
        "*": [
            {"any_keywords": ["GAS"], "category": "Auto", "note": "Gas"},
            {"any_keywords": ["CITY-MARKET", "COSTCO WHSE", "NATURAL GROCERS", "WHOLEFOODS"], "category": "Groceries", "note": "Groceries"},
            {"any_keywords": ["CHIPOTLE"], "category": "Eating Out"},
            {"any_keywords": ["PAWS N CLAWS"], "category": "Gypsy"},
            {"any_keywords": ["MCCREIGHT"], "category": "Health", "note": "Coy and Kiki Dentist"},
            {"any_keywords": ["BRODA"], "category": "Tegan", "note": "Tegan Dentist"},
            {"any_keywords": ["AUTOPAY", "AUTO PAY", "AUTOMATIC PAYMENT", "CRCARDPMT", "VENMO PAYMENT"], "category": "Ignore", "note": "Card payment"},
            {"any_keywords": ["Interest Paid"], "category": "Ignore", "note": "Interest"},
            {"any_keywords": ["Electronic Withdrawal"], "category": "Ignore", "note": "Transfer"},
            {"any_keywords": ["AMAZON PRIME"], "category": "Misc", "note": "Amazon Prime"},
            {"any_keywords": ["AVITA YOGA"], "category": "Misc", "note": "Kiki Yoga"},
            {"any_keywords": ["NYTIMES"], "category": "Misc", "note": "New York Times"},
            {"any_keywords": ["SPOTIFY"], "category": "Misc", "note": "Spotify"},
            {"any_keywords": ["WAL-MART"], "category": "Misc"},
            {"any_keywords": ["T-MOBILE"], "category": "Shelter", "note": "Cell Phone"},
            {"any_keywords": ["SOUTHWES"], "category": "Vacation"},
            {"any_keywords": ["CASH BACK"], "category": "Ignore", "note": "Cashback"},
        ],

        "groups": [
            {
                # optional groups
            },
        ],

        "by_sheet": {
            "CapitalOne_Visa_Quicksilver_7660": [
            ],
            "Chase_Checking_Company_4272": [
            ],
            "Chase_Checking_Farm_5356": [
                {"any_keywords": ["CHEM GRO"], "category": "Ignore"},
            ],
            "Citi_Visa_Costco_0345": [
                {"any_keywords": ["APPLE.COM"], "category": "Misc"},
            ],
            "MVB_Checking_1194": [
                {"any_keywords": ["SCHWAB"], "category": "Ignore", "note": "Transfer from Schwab"},
            ],
            "PayPal_Checking_Coy": [
            ],
            "Schwab_Checking_Joint_4446": [
                {"any_keywords": ["STATE FARM"], "category": "Ignore", "note": "Already in budget overview"},
            ],
            "Venmo_Checking_Coy": [
                {"any_keywords": ["TEGAN"], "category": "Tegan"},
                {"any_keywords": ["JASEY"], "category": "Jasey"},
            ],
            "Venmo_Checking_Kiki": [
                {"any_keywords": ["Christian Van Kirk"], "category": "Misc", "note": "LGMax Storage"},
            ],
            "WellsFargo_Visa_ActiveCash_1758": [
                {"any_keywords": ["SteamboatTennisPB"], "category": "Misc", "note": "Pickleball"},
            ],
        }
    }
}


# ======================================================================
# 2) GENERAL HELPERS
# ======================================================================

def parse_iso_date(s: str) -> date:
    d = pd.to_datetime(s, format="%Y-%m-%d", errors="raise")
    return d.date()


def safe_sheet_name(name: str) -> str:
    invalid = [":", "\\", "/", "?", "*", "[", "]"]
    s = (name or "").strip()
    for ch in invalid:
        s = s.replace(ch, "-")
    if len(s) > 31:
        s = s[:31]
    return s if s else "Sheet"


def normalize_headers(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).strip() for c in df.columns]
    return df


def excel_letters_to_index(s: str) -> int:
    text = s.strip().upper()
    if text == "" or not text.isalpha():
        raise ValueError(f"Invalid column letters: '{s}'")
    idx = 0
    for ch in text:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def coerce_date_series(series: pd.Series) -> pd.Series:
    """
    Parse common bank date formats without Pandas guessing:
      - MM/DD/YYYY (Schwab, many banks)
      - YYYY-MM-DD
      - ISO datetime like 2025-11-14T03:59:27 (Venmo)
    Returns a Series of python datetime.date (or NaT->NaN).
    """
    s = series.fillna("").astype(str).str.strip()

    # Treat empty as missing
    s = s.replace("", pd.NA)

    # 1) Try ISO / mixed ISO first (handles Venmo "2025-11-14T03:59:27")
    dt_iso = pd.to_datetime(s, errors="coerce", format="ISO8601", utc=False)

    # 2) Try MM/DD/YYYY for remaining
    mask = dt_iso.isna() & s.notna()
    if mask.any():
        dt_mdy = pd.to_datetime(s[mask], errors="coerce", format="%m/%d/%Y")
        dt_iso.loc[mask] = dt_mdy

    # 3) Try YYYY-MM-DD for remaining
    mask = dt_iso.isna() & s.notna()
    if mask.any():
        dt_ymd = pd.to_datetime(s[mask], errors="coerce", format="%Y-%m-%d")
        dt_iso.loc[mask] = dt_ymd

    return dt_iso.dt.date


def coerce_money_series(series: pd.Series) -> pd.Series:
    """
    Robust money parsing:
    - Handles ($8.75), -8.75, $1,234.56, unicode minus, blank
    - Returns float series with NaN for non-parsable values
    """
    s = series.fillna("").astype(str)

    # Normalize whitespace (including tabs / NBSP) and strip
    s = s.str.replace("\u00A0", " ", regex=False)  # NBSP
    s = s.str.replace(r"\s+", "", regex=True)      # remove all whitespace

    # Detect negatives via parentheses OR leading minus (including unicode minus)
    s = s.str.replace("\u2212", "-", regex=False)  # unicode minus -> hyphen-minus
    is_paren_neg = s.str.contains(r"^\(.*\)$", regex=True)
    is_minus_neg = s.str.startswith("-")

    # Remove parentheses and leading minus for numeric extraction
    s2 = s.str.replace("(", "", regex=False).str.replace(")", "", regex=False)
    s2 = s2.str.lstrip("-")

    # Remove $ and commas
    s2 = s2.str.replace("$", "", regex=False)
    s2 = s2.str.replace(",", "", regex=False)

    num = pd.to_numeric(s2, errors="coerce")

    neg = is_paren_neg | is_minus_neg
    num.loc[neg & num.notna()] = -num.loc[neg & num.notna()]

    return num



def normalize_text(s: Any) -> str:
    if s is None:
        return ""
    text = str(s)
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def normalize_amount(a: Any) -> float:
    try:
        return round(float(a), 2)
    except Exception:
        return 0.0


def normalize_date_value(d: Any) -> str:
    if d is None:
        return ""
    if isinstance(d, date) and not isinstance(d, datetime):
        return d.isoformat()
    if isinstance(d, datetime):
        return d.date().isoformat()
    try:
        dt = pd.to_datetime(d, errors="coerce")
        if pd.isna(dt):
            return ""
        return dt.date().isoformat()
    except Exception:
        return ""


# ======================================================================
# 3) HEADER SEARCH (Venmo-style)
# ======================================================================

def find_header_row_index(path: str, tokens: List[str], max_lines: int) -> Optional[int]:
    try:
        with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
            for i, line in enumerate(f):
                if i >= max_lines:
                    break
                if all(t in line for t in tokens):
                    return i
    except Exception:
        return None
    return None


# ======================================================================
# 4) CSV READING
# ======================================================================

def read_csv_with_fallbacks(
    path: str,
    *,
    delimiter: Optional[str],
    skiprows: int,
    encoding: Optional[str],
    has_header: bool
) -> pd.DataFrame:
    encodings: List[Optional[str]] = []
    if encoding:
        encodings.append(encoding)
    encodings += ["utf-8-sig", "utf-8", "cp1252"]

    header_arg = 0 if has_header else None

    last_exc: Optional[Exception] = None
    for enc in encodings:
        try:
            if delimiter is None:
                return pd.read_csv(
                    path,
                    dtype=str,
                    sep=None,
                    engine="python",
                    skiprows=skiprows,
                    encoding=enc,
                    header=header_arg
                )
            return pd.read_csv(
                path,
                dtype=str,
                sep=delimiter,
                skiprows=skiprows,
                encoding=enc,
                header=header_arg
            )
        except Exception as e:
            last_exc = e

    raise RuntimeError(f"Failed to read '{path}'. Last error: {last_exc}")


# ======================================================================
# 5) WHOLE-WORD / WHOLE-PHRASE REGEX
# ======================================================================

def token_boundary_wrap(token_pattern: str) -> str:
    return rf"(?<![A-Za-z0-9]){token_pattern}(?![A-Za-z0-9])"


def keyword_to_whole_word_regex(keyword: str) -> str:
    words = [w for w in str(keyword).strip().split() if w]
    if not words:
        return ""

    parts: List[str] = []
    for w in words:
        esc = re.escape(w)
        parts.append(token_boundary_wrap(esc))

    return r"\s+".join(parts)


# ======================================================================
# 6) AUTO CATEGORY + NOTE RULES (global/groups/by_sheet)
# ======================================================================

def get_rules_for_account(account_name: str, rules_cfg: Dict[str, Any]) -> List[Dict[str, Any]]:
    rules: List[Dict[str, Any]] = []

    global_rules = rules_cfg.get("*", [])
    if isinstance(global_rules, list):
        rules.extend(global_rules)

    groups = rules_cfg.get("groups", [])
    if isinstance(groups, list):
        for g in groups:
            if not isinstance(g, dict):
                continue
            sheets = g.get("sheets", [])
            grules = g.get("rules", [])
            if isinstance(sheets, list) and account_name in sheets and isinstance(grules, list):
                rules.extend(grules)

    by_sheet = rules_cfg.get("by_sheet", {})
    if isinstance(by_sheet, dict):
        srules = by_sheet.get(account_name, [])
        if isinstance(srules, list):
            rules.extend(srules)

    return rules


def apply_auto_category_and_note(
    df: pd.DataFrame,
    account_name: str,
    rules_cfg: Dict[str, Any]
) -> pd.DataFrame:
    if df.empty:
        return df

    for col in ["Description", "Category", "Note"]:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].fillna("").astype(str)

    rules = get_rules_for_account(account_name, rules_cfg)
    if not rules:
        return df

    desc = df["Description"].fillna("").astype(str)

    cat_blank = df["Category"].fillna("").astype(str).str.strip().eq("")
    note_blank = df["Note"].fillna("").astype(str).str.strip().eq("")
    if not (cat_blank.any() or note_blank.any()):
        return df

    for rule in rules:
        if not isinstance(rule, dict):
            continue

        category_value = str(rule.get("category", "")).strip()
        note_value = str(rule.get("note", "")).strip()

        any_keywords = rule.get("any_keywords", None)
        all_keywords = rule.get("all_keywords", None)

        match_mask = None

        if isinstance(any_keywords, list) and len(any_keywords) > 0:
            m = False
            for kw in any_keywords:
                pattern = keyword_to_whole_word_regex(kw)
                if not pattern:
                    continue
                m = m | desc.str.contains(pattern, case=False, regex=True, na=False)
            match_mask = m if match_mask is None else (match_mask | m)

        if isinstance(all_keywords, list) and len(all_keywords) > 0:
            m = True
            for kw in all_keywords:
                pattern = keyword_to_whole_word_regex(kw)
                if not pattern:
                    continue
                m = m & desc.str.contains(pattern, case=False, regex=True, na=False)
            match_mask = m if match_mask is None else (match_mask | m)

        if match_mask is None:
            continue

        if category_value != "":
            fill_cat = cat_blank & match_mask
            if fill_cat.any():
                df.loc[fill_cat, "Category"] = category_value
                cat_blank = df["Category"].fillna("").astype(str).str.strip().eq("")

        if note_value != "":
            fill_note = note_blank & match_mask
            if fill_note.any():
                df.loc[fill_note, "Note"] = note_value
                note_blank = df["Note"].fillna("").astype(str).str.strip().eq("")

        if not (cat_blank.any() or note_blank.any()):
            break

    return df


# ======================================================================
# 7) COLUMN MAPPING + AMOUNT DERIVATION
# ======================================================================

def map_columns_from_csv(
    df: pd.DataFrame,
    *,
    account_name: str,
    csv_basename: str,
    has_header: bool,
    match_map: Dict[str, MatchValue],
    needed_columns: List[str]
) -> pd.DataFrame:
    if has_header:
        df = normalize_headers(df)

        if df.shape[1] == 1 and len(df.columns) == 1:
            col0 = str(df.columns[0])
            if "," in col0 or "\t" in col0:
                raise RuntimeError(
                    f"{account_name}: '{csv_basename}' parsed as ONE column.\n"
                    f"Header seen: {col0}\n"
                    f"Fix by setting delimiter to None (auto) or ',' or '\\t' in this account config."
                )

    out: Dict[str, pd.Series] = {}

    for out_col in needed_columns:
        if out_col not in match_map:
            continue

        src = match_map[out_col]

        if has_header:
            if not isinstance(src, str):
                raise ValueError(f"{account_name}: match for '{out_col}' must be a header string when has_header=True")
            if has_header:
                if not isinstance(src, str):
                    raise ValueError(
                        f"{account_name}: match for '{out_col}' must be a header string when has_header=True")

                # Build a normalized-header map so matching is resilient to spacing/case/weird chars
                def norm(h: str) -> str:
                    t = str(h).replace("\u00A0", " ").strip()
                    t = re.sub(r"\s+", " ", t)
                    return t.lower()

                header_map: Dict[str, str] = {}
                for c in df.columns:
                    header_map[norm(c)] = c  # last wins, fine for most exports

                wanted = norm(src)

                if wanted in header_map:
                    out[out_col] = df[header_map[wanted]]
                else:
                    raise KeyError(
                        f"{account_name}: '{csv_basename}' missing header '{src}'. Found: {list(df.columns)}"
                    )

        else:
            if isinstance(src, int):
                src_index = src
            elif isinstance(src, str) and src.strip().isalpha():
                src_index = excel_letters_to_index(src)
            else:
                raise ValueError(
                    f"{account_name}: match for '{out_col}' must be an int or letters like 'A' when has_header=False"
                )

            if src_index < 0 or src_index >= df.shape[1]:
                raise IndexError(
                    f"{account_name}: '{csv_basename}' column index {src_index} out of range (cols={df.shape[1]})"
                )
            out[out_col] = df.iloc[:, src_index]

    mapped = pd.DataFrame(out)

    for col in needed_columns:
        if col not in mapped.columns:
            mapped[col] = ""

    return mapped[needed_columns]


def derive_amount_from_withdraw_deposit(df: pd.DataFrame, withdraw_is_negative: bool) -> pd.DataFrame:
    if "Withdraw" not in df.columns:
        df["Withdraw"] = pd.NA
    if "Deposit" not in df.columns:
        df["Deposit"] = pd.NA

    df["Withdraw"] = coerce_money_series(df["Withdraw"])
    df["Deposit"] = coerce_money_series(df["Deposit"])

    if withdraw_is_negative:
        df["Amount"] = df["Deposit"].fillna(0) + df["Withdraw"].fillna(0)
    else:
        df["Amount"] = df["Deposit"].fillna(0) - df["Withdraw"].fillna(0)

    return df


# ======================================================================
# 8) VENMO ENDING BALANCE HELPERS
# ======================================================================

def get_last_non_null_money_value(df: pd.DataFrame, col_name: str) -> Optional[float]:
    if col_name not in df.columns:
        return None
    vals = coerce_money_series(df[col_name]).dropna()
    if vals.empty:
        return None
    return float(vals.iloc[-1])


def get_max_date_in_column(df: pd.DataFrame, col_name: str) -> Optional[date]:
    if col_name not in df.columns:
        return None
    dts = coerce_date_series(df[col_name]).dropna()
    if dts.empty:
        return None
    return max(dts)


# ======================================================================
# 9) DEDUP KEYS + BATCH DEDUP
# ======================================================================

def make_dedup_key_from_values(dv: Any, desc: Any, amt: Any) -> str:
    """
    DEDUP KEY (NOTE IS IGNORED):
      Date + Description + Amount
    """
    d = normalize_date_value(dv)
    dsc = normalize_text(desc)
    a = normalize_amount(amt)
    return f"{d}|{dsc}|{a:.2f}"


def dedup_transactions_in_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove OVERLAP duplicates across multiple CSV files for the same account
    WITHOUT collapsing legitimate duplicate transactions.

    Problem we are solving:
      - Many bank exports overlap between files, so the SAME transaction can appear in 2 files.
      - Sometimes there are also REAL duplicates (same Date/Description/Amount occurs twice).

    If we simply drop_duplicates on (Date, Description, Amount) we lose real duplicates.

    Strategy used here:
      - Compute the signature key = Date + Description + Amount (Note ignored).
      - For each signature, compute the MAX number of occurrences in ANY ONE source file.
      - Keep up to that max count across the combined batch.

    Why this works well in practice:
      - If file A and file B overlap and both contain the same transaction once:
            max-per-file count = 1, and the combined batch keeps 1.
      - If there are real duplicates and both appear in a file:
            file count = 2, so max-per-file count = 2, and we keep both.

    Limitation:
      - If two real identical transactions are split across different files with NO single file
        containing both occurrences, we cannot distinguish that from overlap using only
        Date/Description/Amount. In that rare case, this rule would keep 1 instead of 2.
        If you ever hit that case, the fix is to include a better unique field (ID, timestamp, etc.)
        into the signature for that specific account.

    Expected input:
      - df includes at least: Date, Description, Amount
      - df SHOULD include: __source_file and __source_order (we add them if missing)

    Returns:
      - DataFrame containing only STANDARD_COLUMNS.
    """
    if df.empty:
        return df[STANDARD_COLUMNS].copy() if all(c in df.columns for c in STANDARD_COLUMNS) else df

    out = df.copy()

    if "__source_file" not in out.columns:
        out["__source_file"] = "(single)"
    if "__source_order" not in out.columns:
        out["__source_order"] = list(range(len(out)))

    out["__dedup_key"] = out.apply(
        lambda r: make_dedup_key_from_values(r.get("Date"), r.get("Description"), r.get("Amount")),
        axis=1
    )

    out = out[out["__dedup_key"] != "||0.00"].copy()

    if out.empty:
        return df[STANDARD_COLUMNS].copy() if all(c in df.columns for c in STANDARD_COLUMNS) else df

    per_file = out.groupby(["__source_file", "__dedup_key"]).size().reset_index(name="cnt")
    max_cnt = per_file.groupby("__dedup_key")["cnt"].max()

    sort_cols: List[str] = []
    if "Date" in out.columns:
        sort_cols.append("Date")
    sort_cols += ["__source_file", "__source_order"]

    out = out.sort_values(by=sort_cols, kind="stable", na_position="last").reset_index(drop=True)

    out["__occ"] = out.groupby("__dedup_key", sort=False).cumcount() + 1
    out["__max"] = out["__dedup_key"].map(max_cnt).fillna(0).astype(int)

    out = out.loc[out["__occ"] <= out["__max"]].copy()

    out = out.drop(
        columns=["__dedup_key", "__occ", "__max", "__source_file", "__source_order"],
        errors="ignore"
    )

    for c in STANDARD_COLUMNS:
        if c not in out.columns:
            out[c] = ""

    return out[STANDARD_COLUMNS].copy()

# ======================================================================
# 10) INGEST ONE ACCOUNT FOLDER (CSV -> standardized DataFrame)
#     Returns: (out_df, ending_balance_value or None)
# ======================================================================

def ingest_account_folder(
    *,
    account_name: str,
    folder_path: str,
    account_cfg: Dict[str, Any],
    date_start: Optional[date],
    date_stop: Optional[date]
) -> Tuple[pd.DataFrame, Optional[float]]:
    csv_files = sorted(glob.glob(os.path.join(folder_path, "*.csv")))
    if not csv_files:
        return (pd.DataFrame(columns=STANDARD_COLUMNS), None)

    has_header = bool(account_cfg.get("has_header", True))
    skiprows = int(account_cfg.get("skiprows", 0))
    delimiter = account_cfg.get("delimiter", None)
    encoding = account_cfg.get("encoding", None)

    header_search_tokens = account_cfg.get("header_search_tokens", None)
    max_scan = int(account_cfg.get("max_header_scan_lines", 200))

    match_map: Dict[str, MatchValue] = dict(account_cfg.get("match", {}))

    needed_columns = list(set(STANDARD_COLUMNS + list(match_map.keys()) + ["Withdraw", "Deposit"]))
    stable_needed: List[str] = []
    for c in STANDARD_COLUMNS:
        if c in needed_columns:
            stable_needed.append(c)
    for c in sorted(needed_columns):
        if c not in stable_needed:
            stable_needed.append(c)
    needed_columns = stable_needed

    frames: List[pd.DataFrame] = []

    ending_balance_candidates: List[Tuple[Optional[date], float]] = []
    want_ending_balance = bool(account_cfg.get("place_ending_balance_on_last_transaction_row", False))
    eb_header = str(account_cfg.get("ending_balance_header", "")).strip()
    dt_header = str(account_cfg.get("date_header_for_balance_pick", "")).strip()

    for csv_path in csv_files:
        csv_basename = os.path.basename(csv_path)

        effective_skiprows = skiprows
        if has_header and header_search_tokens:
            header_idx = find_header_row_index(csv_path, list(header_search_tokens), max_scan)
            if header_idx is None:
                raise RuntimeError(
                    f"{account_name}: Could not find header row in '{csv_basename}' using tokens {header_search_tokens}"
                )
            effective_skiprows = header_idx

        raw = read_csv_with_fallbacks(
            csv_path,
            delimiter=delimiter,
            skiprows=effective_skiprows,
            encoding=encoding,
            has_header=has_header
        )

        if has_header:
            raw = normalize_headers(raw)

        if want_ending_balance and has_header and eb_header != "":
            eb_val = get_last_non_null_money_value(raw, eb_header)
            if eb_val is not None:
                max_dt = get_max_date_in_column(raw, dt_header) if dt_header else None
                ending_balance_candidates.append((max_dt, eb_val))

        mapped = map_columns_from_csv(
            raw,
            account_name=account_name,
            csv_basename=csv_basename,
            has_header=has_header,
            match_map=match_map,
            needed_columns=needed_columns
        )

        if "Date" in mapped.columns:
            mapped["Date"] = coerce_date_series(mapped["Date"])

        for col in ["Amount", "Balance", "Withdraw", "Deposit"]:
            if col in mapped.columns:
                mapped[col] = coerce_money_series(mapped[col])

        if bool(account_cfg.get("derive_amount_from_withdraw_deposit", False)):
            mapped = derive_amount_from_withdraw_deposit(
                mapped,
                withdraw_is_negative=bool(account_cfg.get("withdraw_is_negative", False))
            )

        for col in ["Description", "Note", "Category"]:
            if col not in mapped.columns:
                mapped[col] = ""
            mapped[col] = mapped[col].fillna("").astype(str)

        mapped = apply_auto_category_and_note(
            mapped,
            account_name,
            CONFIG.get("auto_category_rules", {})
        )

        # Drop non-transaction rows (no date)
        if "Date" in mapped.columns:
            mapped = mapped.loc[mapped["Date"].notna()]

        # Apply date-range filter (inclusive)
        if date_start is not None and date_stop is not None:
            dates = pd.to_datetime(mapped["Date"], errors="coerce").dt.date  # <- python date objects
            mapped = mapped.loc[(dates >= date_start) & (dates <= date_stop)]

        mapped["__source_file"] = csv_basename
        mapped["__source_order"] = list(range(len(mapped)))

        frames.append(mapped)

    combined = pd.concat(frames, ignore_index=True)

    if "Date" in combined.columns:
        combined = combined.sort_values(by=["Date"], kind="stable", na_position="last").reset_index(drop=True)

    for c in STANDARD_COLUMNS:
        if c not in combined.columns:
            combined[c] = ""

    # Keep source columns so we can dedup overlap across files without collapsing real duplicates
    keep_cols = list(STANDARD_COLUMNS)
    for extra in ["__source_file", "__source_order"]:
        if extra in combined.columns:
            keep_cols.append(extra)

    out = combined[keep_cols].copy()

    out["Description"] = out["Description"].fillna("").astype(str).str.strip()
    out["Note"] = out["Note"].fillna("").astype(str).str.strip()
    out["Category"] = out["Category"].fillna("").astype(str).str.strip()

    out = dedup_transactions_in_dataframe(out)

    ending_balance_value: Optional[float] = None
    if ending_balance_candidates:
        ending_balance_candidates.sort(key=lambda x: (x[0] is None, x[0]))
        ending_balance_value = float(ending_balance_candidates[-1][1])

    return (out, ending_balance_value)


# ======================================================================
# 11) EXCEL HELPERS: find last transaction row, dropdown, formulas, append, formatting
# ======================================================================

COLUMN_WIDTHS: List[float] = [15, 50, 50, 15, 15, 15]  # A..F
COLUMN_ALIGNMENTS: List[str] = ["right", "left", "left", "left", "right", "right"]  # A..F


def ensure_categories_sheet_last(wb, categories: List[str]) -> Tuple[str, int]:
    sheet_name = safe_sheet_name("Categories")
    last_row = 1 + len(categories)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)

    ws["A1"].value = "Category"
    for i, cat in enumerate(categories, start=2):
        ws[f"A{i}"].value = cat

    ws_obj = wb[sheet_name]
    wb._sheets.remove(ws_obj)
    wb._sheets.append(ws_obj)

    return sheet_name, last_row


def ensure_sheet_with_headers(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 1:
            ws.append(STANDARD_COLUMNS)
        else:
            headers = [ws.cell(row=1, column=i).value for i in range(1, len(STANDARD_COLUMNS) + 1)]
            headers_norm = [str(h).strip() if h is not None else "" for h in headers]
            if headers_norm != STANDARD_COLUMNS:
                ws.delete_rows(1, 1)
                ws.insert_rows(1)
                for i, h in enumerate(STANDARD_COLUMNS, start=1):
                    ws.cell(row=1, column=i).value = h
        return ws

    ws = wb.create_sheet(sheet_name)
    ws.append(STANDARD_COLUMNS)
    return ws


def find_last_transaction_row(ws) -> int:
    """
    Last row that actually contains a transaction (not formatting).
    Looks for any value in Date (A), Description (B), or Amount (E).
    """
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=1).value not in (None, ""):
            return r
        if ws.cell(row=r, column=2).value not in (None, ""):
            return r
        if ws.cell(row=r, column=5).value not in (None, ""):
            return r
    return 1


def reset_category_dropdown(ws, categories_sheet_name: str, categories_last_row: int, min_rows: int) -> None:
    if categories_last_row < 2:
        return

    ws.data_validations.dataValidation = []

    formula = f"='{categories_sheet_name}'!$A$2:$A${categories_last_row}"
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.errorTitle = "Invalid Category"
    dv.error = "Please select a Category from the dropdown list."
    ws.add_data_validation(dv)

    category_col_1based = STANDARD_COLUMNS.index("Category") + 1
    col_letter = get_column_letter(category_col_1based)

    end_row = max(find_last_transaction_row(ws), min_rows, 2)
    dv.add(f"{col_letter}2:{col_letter}{end_row}")


def apply_running_balance_formula_excluding_ignore(ws) -> None:
    last_row = find_last_transaction_row(ws)
    if last_row < 2:
        return

    cat_col = "D"
    amt_col = "E"
    bal_col = "F"

    ws[f"{bal_col}2"].value = f'=IF({amt_col}2="","",IF(ISNUMBER(SEARCH("Ignore",{cat_col}2)),0,{amt_col}2))'
    ws[f"{bal_col}2"].number_format = "0.00"

    for r in range(3, last_row + 1):
        ws[f"{bal_col}{r}"].value = (
            f'=IF({amt_col}{r}="","",{bal_col}{r-1}+IF(ISNUMBER(SEARCH("Ignore",{cat_col}{r})),0,{amt_col}{r}))'
        )
        ws[f"{bal_col}{r}"].number_format = "0.00"


def set_ending_balance_on_last_transaction_row(ws, value: float) -> None:
    r = find_last_transaction_row(ws)
    if r < 2:
        return
    ws.cell(row=r, column=6).value = float(value)  # F
    ws.cell(row=r, column=6).number_format = "0.00"


def get_existing_signature_counts(ws) -> Dict[str, int]:
    """
    Count how many times each (Date + Description + Amount) signature appears in the EXISTING sheet.
    Note is intentionally ignored.
    """
    counts: Dict[str, int] = {}
    last_row = find_last_transaction_row(ws)
    if last_row < 2:
        return counts

    for r in range(2, last_row + 1):
        dv = ws.cell(row=r, column=1).value  # Date
        desc = ws.cell(row=r, column=2).value  # Description
        amt = ws.cell(row=r, column=5).value  # Amount

        key = make_dedup_key_from_values(dv, desc, amt)
        if key != "||0.00":
            counts[key] = counts.get(key, 0) + 1

    return counts


def choose_rows_to_append_by_counts(df_new: pd.DataFrame, existing_counts: Dict[str, int]) -> pd.DataFrame:
    """
    Append-safe filtering that preserves legitimate duplicates.

    For each signature:
      - Excel already contains N occurrences
      - Newly ingested df has M occurrences
      - Append only the (M - N) additional occurrences

    This prevents:
      - Re-appending the same rows when you run the script again
      - Collapsing real duplicates that have identical Date/Description/Amount
    """
    if df_new.empty:
        return df_new

    df = df_new.copy()

    df["__sig"] = df.apply(
        lambda r: make_dedup_key_from_values(r.get("Date"), r.get("Description"), r.get("Amount")),
        axis=1
    )

    df = df[df["__sig"] != "||0.00"].copy()
    if df.empty:
        return df_new.iloc[0:0].copy()

    sort_cols: List[str] = []
    if "Date" in df.columns:
        sort_cols.append("Date")
    sort_cols += ["Description", "Amount"]

    # If Balance exists, include it to make ordering more deterministic
    if "Balance" in df.columns and df["Balance"].notna().any():
        sort_cols.append("Balance")

    df = df.sort_values(by=sort_cols, kind="stable", na_position="last").reset_index(drop=True)

    df["__occ"] = df.groupby("__sig", sort=False).cumcount() + 1
    df["__existing"] = df["__sig"].map(existing_counts).fillna(0).astype(int)

    out = df.loc[df["__occ"] > df["__existing"]].copy()
    out = out.drop(columns=["__sig", "__occ", "__existing"], errors="ignore")

    return out


def append_new_rows(ws, df_new: pd.DataFrame) -> int:
    if df_new.empty:
        return 0

    start_row = find_last_transaction_row(ws) + 1
    appended = 0

    for _, row in df_new.iterrows():
        excel_row = start_row + appended

        for col_idx, col_name in enumerate(STANDARD_COLUMNS, start=1):
            v = row.get(col_name, "")

            if col_name == "Date":
                if isinstance(v, date) and not isinstance(v, datetime):
                    ws.cell(row=excel_row, column=col_idx).value = v
                else:
                    dt = pd.to_datetime(v, errors="coerce")
                    ws.cell(row=excel_row, column=col_idx).value = dt.date() if not pd.isna(dt) else ""
            elif col_name in ["Amount", "Balance"]:
                try:
                    fv = float(v)
                    if pd.isna(fv):
                        ws.cell(row=excel_row, column=col_idx).value = ""
                    else:
                        ws.cell(row=excel_row, column=col_idx).value = fv
                        ws.cell(row=excel_row, column=col_idx).number_format = "0.00"
                except Exception:
                    ws.cell(row=excel_row, column=col_idx).value = ""
            else:
                ws.cell(row=excel_row, column=col_idx).value = "" if v is None else str(v)

        appended += 1

    return appended


def apply_sheet_formatting(ws) -> None:
    """
    Apply:
    - Column widths (A..F)
    - Alignment per column (A..F)
    - Wrap text for all cells through the last REAL transaction row
    """
    for i, w in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    end_row = max(find_last_transaction_row(ws), 1)
    end_col = len(STANDARD_COLUMNS)

    for col_idx in range(1, end_col + 1):
        horiz = COLUMN_ALIGNMENTS[col_idx - 1]
        align = Alignment(horizontal=horiz, wrap_text=True)

        for row_idx in range(1, end_row + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = align


# ======================================================================
# 12) MAIN
# ======================================================================

def main() -> None:
    workbook_name = str(CONFIG.get("workbook_name", "")).strip() or "Workbook"

    date_start: Optional[date] = None
    date_stop: Optional[date] = None
    dr = CONFIG.get("date_range") or {}
    if dr.get("start") and dr.get("stop"):
        date_start = parse_iso_date(dr["start"])
        date_stop = parse_iso_date(dr["stop"])

    script_dir = os.path.dirname(os.path.abspath(__file__))

    budget_root = os.path.join(script_dir, CONFIG["budget_accounts_folder"])
    if not os.path.isdir(budget_root):
        raise FileNotFoundError(f"Budget accounts folder not found: {budget_root}")

    output_path = os.path.join(script_dir, f"{workbook_name}.xlsx")
    workbook_exists = os.path.isfile(output_path)

    if workbook_exists:
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        if wb.sheetnames:
            del wb[wb.sheetnames[0]]

    categories: List[str] = CONFIG.get("categories") or []
    categories_sheet_name, categories_last_row = ensure_categories_sheet_last(wb, categories)

    accounts_by_name: Dict[str, Dict[str, Any]] = CONFIG.get("accounts_by_name", {})
    skip_unknown = bool(CONFIG.get("skip_unknown_account_folders", True))
    min_rows = int(CONFIG.get("category_validation_min_rows", 1000))

    for entry in sorted(os.listdir(budget_root)):
        folder_path = os.path.join(budget_root, entry)
        if not os.path.isdir(folder_path):
            continue

        account_name = entry
        if account_name not in accounts_by_name:
            msg = f"Found account folder with no config: '{account_name}'"
            if skip_unknown:
                print(f"WARNING: {msg} (skipping)")
                continue
            raise KeyError(msg)

        account_cfg = accounts_by_name[account_name]
        sheet_name = safe_sheet_name(account_cfg.get("sheet_name") or account_name)
        ws = ensure_sheet_with_headers(wb, sheet_name)

        existing_counts = get_existing_signature_counts(ws)

        df_all, ending_balance_value = ingest_account_folder(
            account_name=account_name,
            folder_path=folder_path,
            account_cfg=account_cfg,
            date_start=date_start,
            date_stop=date_stop
        )

        df_to_append = choose_rows_to_append_by_counts(df_all, existing_counts)

        appended = append_new_rows(ws, df_to_append)
        if appended > 0:
            print(f"{sheet_name}: appended {appended} new row(s)")
        else:
            print(f"{sheet_name}: no new rows to append")

        reset_category_dropdown(ws, categories_sheet_name, categories_last_row, min_rows)

        balance_mode = str(account_cfg.get("balance_mode", "")).strip().lower()
        if balance_mode == "running_balance_excluding_ignore":
            apply_running_balance_formula_excluding_ignore(ws)

        if ending_balance_value is not None and bool(account_cfg.get("place_ending_balance_on_last_transaction_row", False)):
            set_ending_balance_on_last_transaction_row(ws, ending_balance_value)

        apply_sheet_formatting(ws)

    ensure_categories_sheet_last(wb, categories)

    wb.save(output_path)
    print(f"Saved workbook: {output_path}")


if __name__ == "__main__":
    main()
